from __future__ import annotations

from pathlib import Path
from typing import Any


def _parse_sheet1_performance(rows: list[tuple]) -> dict[str, Any]:
    key_map = {
        "post url": "post_url",
        "post date": "post_date",
        "post publish time": "post_publish_time",
        "impressions": ("impressions", int),
        "members reached": ("members_reached", int),
        "profile viewers from this post": ("profile_viewers", int),
        "followers gained from this post": ("followers_gained", int),
        "reactions": ("reactions", int),
        "comments": ("comments", int),
        "reposts": ("reposts", int),
        "saves": ("saves", int),
        "sends on linkedin": ("sends", int),
        "top job title": "top_job_title",
        "top location": "top_location",
        "top industry": "top_industry",
    }
    result: dict[str, Any] = {}
    for row in rows:
        if not row or len(row) < 2:
            continue
        key = str(row[0] or "").strip().lower()
        value = row[1]
        if value is None:
            continue
        if key in key_map:
            mapped = key_map[key]
            if isinstance(mapped, tuple):
                field_name, converter = mapped
                try:
                    result[field_name] = converter(value)
                except (ValueError, TypeError):
                    result[field_name] = None
            else:
                result[mapped] = str(value or "").strip()
    return result


def _parse_sheet2_demographics(rows: list[tuple]) -> list[dict[str, Any]]:
    demographics: list[dict[str, Any]] = []
    for row in rows:
        if not row or len(row) < 3:
            continue
        category = str(row[0] or "").strip()
        value = str(row[1] or "").strip()
        pct_raw = row[2]
        if category.lower() in ("category", ""):
            continue
        if not category or not value:
            continue
        try:
            if isinstance(pct_raw, str):
                pct_raw = pct_raw.strip().rstrip('%')
            pct_float = float(pct_raw)
            if pct_float <= 1.0:
                percentage = round(pct_float * 100, 1)
            else:
                percentage = round(pct_float, 1)
        except (ValueError, TypeError):
            percentage = None
        demographics.append({
            "category": category,
            "value": value,
            "percentage": percentage,
        })
    return demographics


def parse_linkedin_export_xlsx(filepath: str) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    path = Path(filepath)
    if not path.is_file():
        return {}

    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return {}

    sheet_names = wb.sheetnames
    result: dict[str, Any] = {"demographics": []}

    if sheet_names:
        ws1 = wb[sheet_names[0]]
        rows1 = list(ws1.iter_rows(min_row=1, values_only=True))
        performance = _parse_sheet1_performance(rows1)
        result.update(performance)

    if len(sheet_names) > 1:
        ws2 = wb[sheet_names[1]]
        rows2 = list(ws2.iter_rows(min_row=1, values_only=True))
        result["demographics"] = _parse_sheet2_demographics(rows2)

    wb.close()
    return result